"""
통합 실행기 - 구글 스프레드시트 + 텔레그램 연동
  - 구글시트에서 "대기" 계정 목록 가져오기
  - 계정별 강의+퀴즈 통합 자동화
  - 진행 상태를 구글시트에 실시간 업데이트
  - 텔레그램 알림 (스크린샷, 수료증 PDF, 에러)
  - 텔레그램 원격 제어 (/status, /restart)
  - 엑셀 파일 폴백 지원 (구글시트 미설정 시)
"""

import sys
import time
import tkinter as tk
from tkinter import filedialog

from openpyxl import load_workbook

import config
import gsheet
import telegram_bot
import certificate
from main import create_driver, login, run_lectures, get_lecture_courses
from telegram_bot import RunnerState, TelegramController


MAX_RESTART = 3  # 셀레니움 완전 재시작 최대 횟수


# ── 엑셀 폴백 (구글시트 미설정 시) ──

def select_excel_file():
    """tkinter 파일 다이얼로그로 .xlsx 파일 선택"""
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="계정 엑셀 파일 선택",
        filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")],
    )
    root.destroy()
    return file_path


def read_accounts_from_excel(file_path):
    """엑셀에서 ID/PW 목록 읽기 (첫 행 헤더, 이후 데이터)"""
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    accounts = []
    for row in rows:
        if row and len(row) >= 2 and row[0] and row[1]:
            acc = {
                "user_id": str(row[0]).strip(),
                "user_pw": str(row[1]).strip(),
            }
            # C열에 telegram_chat_id가 있으면 추가
            if len(row) >= 3 and row[2]:
                acc["telegram_chat_id"] = str(row[2]).strip()
            accounts.append(acc)
    return accounts


# ── 스크린샷 ──

def take_screenshot(driver):
    """Selenium 스크린샷 → PNG bytes 반환"""
    try:
        return driver.get_screenshot_as_png()
    except Exception as e:
        print(f"  [스크린샷] 촬영 실패: {e}")
        return None


def send_status_screenshot(driver, user_id, caption, chat_id=None):
    """강의실 화면 스크린샷 → 텔레그램 전송"""
    screenshot = take_screenshot(driver)
    if screenshot:
        telegram_bot.send_photo(screenshot, caption=caption, chat_id=chat_id)


# ── 수료증 처리 ──

def process_certificates(driver, user_id, chat_id=None):
    """마이페이지에서 수료증 PDF 추출 → 텔레그램 전송"""
    print(f"\n  [수료증] 수료증 확인 중...")
    driver.get(config.MYPAGE_URL)
    time.sleep(3)

    certs = certificate.extract_cert_js_from_mypage(driver)
    if not certs:
        print(f"  [수료증] 수료증 버튼을 찾지 못했습니다.")
        return

    print(f"  [수료증] {len(certs)}개 수료증 발견")
    for cert_info in certs:
        title = cert_info.get("title", "")
        cert_js = cert_info.get("cert_js", "")
        if cert_js:
            print(f"  [수료증] {title or '과정'} PDF 추출 중...")
            ok = certificate.download_and_send_certificate(
                driver, cert_js, user_id, chat_id=chat_id,
            )
            if not ok:
                print(f"  [수료증] v1 실패 → v2(별도 브라우저) 방식으로 재시도...")
                certificate.download_and_send_certificate_v2(
                    driver, cert_js, user_id, chat_id=chat_id,
                )
            time.sleep(2)


# ── 계정별 실행 ──

def run_for_account(user_id, user_pw, state, chat_id=None):
    """단일 계정에 대해 강의+퀴즈 통합 실행"""

    # 구글시트 상태: 진행중
    gsheet.update_status(user_id, "진행중", "시작")
    telegram_bot.send_message(f"[{user_id}] 교육 수강 시작", chat_id=chat_id)

    for attempt in range(1, MAX_RESTART + 1):
        driver = create_driver()

        try:
            # 로그인
            if not login(driver, user_id, user_pw):
                print(f"  → [{user_id}] 로그인 실패!")
                gsheet.update_status(user_id, "실패", "로그인 실패")
                telegram_bot.send_message(f"[{user_id}] 로그인 실패", chat_id=chat_id)
                state.mark_failed(user_id)
                return

            done = run_lectures(driver, user_id, user_pw)

            if done:
                # 완료 → 스크린샷 + 구글시트 업데이트
                gsheet.update_status(user_id, "수료완료", "모든 과정 완료")

                # 강의실 화면 스크린샷 전송
                driver.get(config.MYPAGE_URL)
                time.sleep(3)
                send_status_screenshot(
                    driver, user_id,
                    caption=f"[{user_id}] 모든 과정 수료 완료!",
                    chat_id=chat_id,
                )

                # 수료증 PDF 처리
                process_certificates(driver, user_id, chat_id=chat_id)

                telegram_bot.send_message(f"[{user_id}] 수료 완료!", chat_id=chat_id)
                state.mark_completed()
                return

            # run_lectures가 False 반환 (미완료 과정 남음)
            print(f"\n  [재시작 {attempt}/{MAX_RESTART}] 미완료 과정 남음 → 셀레니움 재시작")
            gsheet.update_status(user_id, "진행중", f"재시작 {attempt}/{MAX_RESTART}")

        except KeyboardInterrupt:
            raise
        except Exception as e:
            print(f"\n  → [{user_id}] 오류 발생: {e}")
            gsheet.update_status(user_id, "진행중", f"오류 후 재시작 {attempt}/{MAX_RESTART}")

            # 에러 스크린샷 전송
            try:
                send_status_screenshot(
                    driver, user_id,
                    caption=f"[{user_id}] 오류: {str(e)[:200]}",
                    chat_id=chat_id,
                )
            except Exception:
                pass

        finally:
            try:
                driver.quit()
            except Exception:
                pass

        time.sleep(5)

    # 최대 재시작 초과
    gsheet.update_status(user_id, "실패", f"최대 재시작({MAX_RESTART}회) 초과")
    telegram_bot.send_message(
        f"[{user_id}] 최대 재시작 초과. 실패 처리됨.",
        chat_id=chat_id,
    )
    state.mark_failed(user_id)


# ── 메인 ──

def main():
    print("=" * 60)
    print("  통합 실행기 - 구글시트 + 텔레그램 연동")
    print("=" * 60)

    # 1. 계정 목록 가져오기 (구글시트 우선, 없으면 엑셀 폴백)
    accounts = []

    if config.GSHEET_WEB_APP_URL:
        print("\n[구글시트] 대기 계정 조회 중...")
        accounts = gsheet.fetch_pending_accounts()

    if not accounts:
        if config.GSHEET_WEB_APP_URL:
            print("[구글시트] 대기 계정 없음. 엑셀 파일로 폴백합니다.")
        print("\n엑셀 파일을 선택하세요...")
        file_path = select_excel_file()

        if not file_path:
            print("파일을 선택하지 않았습니다. 종료합니다.")
            sys.exit(0)

        print(f"선택된 파일: {file_path}")
        accounts = read_accounts_from_excel(file_path)

    if not accounts:
        print("계정을 찾지 못했습니다. 종료합니다.")
        sys.exit(1)

    print(f"\n총 {len(accounts)}개 계정:")
    for i, acc in enumerate(accounts, 1):
        masked_pw = acc["user_pw"][:2] + "*" * (len(acc["user_pw"]) - 2)
        chat_info = f" (chat: {acc['telegram_chat_id']})" if acc.get("telegram_chat_id") else ""
        print(f"  {i}. {acc['user_id']} / {masked_pw}{chat_info}")

    # 2. 상태 객체 + 텔레그램 원격 제어 시작
    state = RunnerState()
    state.total_accounts = len(accounts)

    controller = TelegramController(state, gsheet_module=gsheet)
    controller.start()

    # 전체 시작 알림
    telegram_bot.send_message(f"자동 수강 시작: {len(accounts)}개 계정")

    # 3. 계정별 순차 실행
    try:
        for idx, acc in enumerate(accounts, 1):
            user_id = acc["user_id"]
            user_pw = acc["user_pw"]
            chat_id = acc.get("telegram_chat_id") or None

            print(f"\n{'#' * 60}")
            print(f"  계정 {idx}/{len(accounts)}: {user_id}")
            print(f"{'#' * 60}")

            state.set_current(user_id)
            run_for_account(user_id, user_pw, state, chat_id=chat_id)

            print(f"\n  → [{user_id}] 처리 완료!")
            if idx < len(accounts):
                print("  → 5초 후 다음 계정으로 넘어갑니다...")
                time.sleep(5)

    except KeyboardInterrupt:
        print("\n\n사용자가 중단했습니다. (Ctrl+C)")
        telegram_bot.send_message("사용자가 수동 중단했습니다. (Ctrl+C)")

    # 4. 완료
    controller.stop()

    with state.lock:
        completed = state.completed_accounts
        failed = list(state.failed_accounts)

    summary = f"전체 처리 완료: {completed}/{len(accounts)}개"
    if failed:
        summary += f"\n실패: {', '.join(failed)}"
    telegram_bot.send_message(summary)

    print(f"\n{'=' * 60}")
    print(f"  {summary}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
